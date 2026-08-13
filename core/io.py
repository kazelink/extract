from __future__ import annotations

import csv
import logging
import os
import tempfile
import threading
import time
from datetime import date, datetime, time as dtime
from typing import Tuple

import pandas as pd

try:
    import xlwt
except ImportError:
    xlwt = None


logger = logging.getLogger(__name__)

_ENCODINGS = ["utf-8-sig", "utf-8", "gb18030", "gbk"]
_MODERN_EXCEL_EXTS = {".xlsx", ".xlsm"}
_LEGACY_EXCEL_EXTS = {".xls"}
_EXCEL_EXTS = _MODERN_EXCEL_EXTS | _LEGACY_EXCEL_EXTS
_READ_EXCEPTIONS = (ValueError, OSError)
_WRITE_EXCEPTIONS = (OSError, ValueError, ImportError)
_XLS_MAX_ROWS = 65536
_XLS_MAX_COLS = 256

_TMP_PREFIX = ".tmp_"
_STALE_TMP_AGE_SEC = 60
_WRITE_LOCK = threading.Lock()

_TXT_DELIMITERS = ",\t;|"


def _remove_quietly(path: str) -> None:
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        logger.warning("Failed to remove temp data file: %s", path, exc_info=True)


def sweep_stale_temp_files(path: str) -> int:
    ext = os.path.splitext(path)[1].lower()
    if not ext:
        return 0
    dir_name = os.path.dirname(path) or "."
    try:
        names = os.listdir(dir_name)
    except OSError:
        return 0
    now = time.time()
    removed = 0
    for name in names:
        if not (name.startswith(_TMP_PREFIX) and name.lower().endswith(ext)):
            continue
        stale_path = os.path.join(dir_name, name)
        try:
            if now - os.path.getmtime(stale_path) < _STALE_TMP_AGE_SEC:
                continue
        except OSError:
            continue
        _remove_quietly(stale_path)
        removed += 1
    return removed


def read_data_file(path: str) -> Tuple[pd.DataFrame, str]:
    ext = os.path.splitext(path)[1].lower()
    if ext in _EXCEL_EXTS:
        try:
            return pd.read_excel(path), "utf-8-sig"
        except (OSError, ValueError):
            raise
        except Exception as exc:
            raise ValueError(f"无法读取 Excel 文件：{exc}") from exc

    last_error: Exception | None = None
    for enc in _ENCODINGS:
        try:
            if ext == ".tsv":
                return pd.read_csv(path, sep="\t", encoding=enc, dtype=str), enc
            if ext == ".txt":
                return _read_txt_file(path, enc), enc
            return pd.read_csv(path, encoding=enc, dtype=str), enc
        except _READ_EXCEPTIONS as exc:
            last_error = exc
            logger.debug("Failed to read %s with encoding %s", path, enc, exc_info=True)
        except Exception as exc:
            last_error = exc
            logger.warning("Unexpected error reading %s with encoding %s", path, enc, exc_info=True)

    detail = f"：{last_error}" if last_error else ""
    raise ValueError(f"无法读取数据文件：{path}{detail}") from last_error


def _read_txt_file(path: str, encoding: str) -> pd.DataFrame:
    delimiter = _sniff_txt_delimiter(path, encoding)
    if delimiter and _txt_fields_consistent(path, encoding, delimiter):
        try:
            df = pd.read_csv(path, sep=delimiter, encoding=encoding, dtype=str)
            if len(df.columns) > 1 and len(df) >= 1:
                return df
        except Exception:
            pass
    return _read_lines_as_single_column(path, encoding)


def _sniff_txt_delimiter(path: str, encoding: str) -> str | None:
    try:
        with open(path, encoding=encoding, newline="") as f:
            sample = f.read(8192)
        if not sample.strip():
            return None
        dialect = csv.Sniffer().sniff(sample, delimiters=_TXT_DELIMITERS)
        return dialect.delimiter
    except (OSError, UnicodeDecodeError, csv.Error):
        return None


def _txt_fields_consistent(path: str, encoding: str, delimiter: str) -> bool:
    try:
        with open(path, encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            counts: set[int] = set()
            for row in reader:
                if not any(cell.strip() for cell in row):
                    continue
                counts.add(len(row))
                if len(counts) > 1:
                    return False
        return bool(counts)
    except (OSError, UnicodeDecodeError, csv.Error):
        return False


def _read_lines_as_single_column(path: str, encoding: str) -> pd.DataFrame:
    with open(path, encoding=encoding) as f:
        lines = [line.rstrip("\r\n") for line in f]
    return pd.DataFrame({"text": lines})


def _coerce_xls_value(
    value: object,
    date_style: object,
    datetime_style: object,
) -> tuple[object, object | None]:
    if isinstance(value, (list, tuple, dict, set)):
        return str(value), None
    if value is None:
        return "", None
    try:
        if pd.isna(value):
            return "", None
    except (TypeError, ValueError):
        return str(value), None
    if isinstance(value, pd.Timestamp):
        if value.time() == dtime(0, 0):
            return value.to_pydatetime(), date_style
        return value.to_pydatetime(), datetime_style
    if isinstance(value, datetime):
        if value.time() == dtime(0, 0):
            return value, date_style
        return value, datetime_style
    if isinstance(value, date):
        return value, date_style
    return value, None


def _write_legacy_xls(df: pd.DataFrame, path: str, source_path: str | None = None) -> None:
    if xlwt is None:
        raise ImportError("Writing .xls files requires xlwt.")
    if len(df.columns) > _XLS_MAX_COLS:
        raise ValueError(".xls supports at most 256 columns.")
    if len(df) + 1 > _XLS_MAX_ROWS:
        raise ValueError(".xls supports at most 65535 data rows.")

    workbook = xlwt.Workbook()
    header_style = xlwt.easyxf("font: bold on;")
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    datetime_style = xlwt.easyxf(num_format_str="YYYY-MM-DD HH:MM:SS")

    def write_df_sheet(sheet_name: str) -> None:
        sheet = workbook.add_sheet(sheet_name)
        for col_idx, col_name in enumerate(df.columns):
            sheet.write(0, col_idx, str(col_name), header_style)
        for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=1):
            for col_idx, value in enumerate(row):
                cell_value, style = _coerce_xls_value(value, date_style, datetime_style)
                if style is None:
                    sheet.write(row_idx, col_idx, cell_value)
                else:
                    sheet.write(row_idx, col_idx, cell_value, style)

    first_sheet_name = "Sheet1"
    first_sheet_added = False
    if source_path and os.path.exists(source_path):
        try:
            import xlrd

            book = xlrd.open_workbook(source_path)
            for sheet_idx, src in enumerate(book.sheets()):
                if sheet_idx == 0:
                    first_sheet_name = src.name[:31] or "Sheet1"
                    write_df_sheet(first_sheet_name)
                    first_sheet_added = True
                    continue
                dst = workbook.add_sheet(src.name[:31] or f"Sheet{sheet_idx + 1}")
                for r in range(src.nrows):
                    for c in range(src.ncols):
                        cell = src.cell(r, c)
                        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                            continue
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            value = xlrd.xldate_as_datetime(cell.value, book.datemode)
                            dst.write(r, c, value, datetime_style)
                            continue
                        if cell.ctype == xlrd.XL_CELL_TEXT:
                            value = str(cell.value)
                        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            value = bool(cell.value)
                        elif cell.ctype == xlrd.XL_CELL_ERROR:
                            value = str(cell.value)
                        else:
                            value = cell.value
                        dst.write(r, c, value)
        except Exception:
            logger.warning(
                "Failed to preserve other sheets for %s; writing a single-sheet file.",
                source_path,
                exc_info=True,
            )
        if not first_sheet_added:
            write_df_sheet(first_sheet_name)
    else:
        write_df_sheet(first_sheet_name)

    workbook.save(path)


def _to_excel_cell(value: object) -> object:
    if isinstance(value, (list, tuple, dict, set)):
        return str(value)
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        return str(value)
    if isinstance(value, (int, float, str, bool)) or hasattr(value, "isoformat"):
        return value
    return str(value)


def _fill_excel_sheet(ws, df: pd.DataFrame) -> None:
    ws.append([str(c) for c in df.columns])
    for row in df.itertuples(index=False, name=None):
        ws.append([_to_excel_cell(value) for value in row])


def _write_excel_preserving_sheets(df: pd.DataFrame, source_path: str, tmp_path: str, ext: str) -> None:
    try:
        import openpyxl
    except ImportError:
        df.to_excel(tmp_path, index=False)
        return

    if not os.path.exists(source_path):
        df.to_excel(tmp_path, index=False)
        return
    try:
        wb = openpyxl.load_workbook(source_path, keep_vba=(ext == ".xlsm"))
    except Exception:
        logger.warning(
            "Failed to load workbook %s; rebuilding a single-sheet workbook.",
            source_path,
            exc_info=True,
        )
        df.to_excel(tmp_path, index=False)
        return

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    first_name = wb.sheetnames[0]
    try:
        del wb[first_name]
        ws = wb.create_sheet(title=first_name, index=0)
        _fill_excel_sheet(ws, df)
        wb.active = 0
        wb.save(tmp_path)
    finally:
        vba_archive = getattr(wb, "vba_archive", None)
        try:
            wb.close()
        except Exception:
            logger.warning("Failed to close workbook %s", source_path, exc_info=True)
        if vba_archive is not None:
            try:
                vba_archive.close()
            except Exception:
                logger.warning("Failed to close VBA archive %s", source_path, exc_info=True)


def write_data_file(df: pd.DataFrame, path: str, encoding: str) -> None:
    ext = os.path.splitext(path)[1].lower()
    dir_name = os.path.dirname(path) or "."
    with _WRITE_LOCK:
        tmp_path: str | None = None
        try:
            fd, tmp_path = tempfile.mkstemp(dir=dir_name, prefix=_TMP_PREFIX, suffix=ext)
            os.close(fd)
            if ext in _MODERN_EXCEL_EXTS:
                _write_excel_preserving_sheets(df, path, tmp_path, ext)
            elif ext in _LEGACY_EXCEL_EXTS:
                _write_legacy_xls(df, tmp_path, source_path=path)
            else:
                df.to_csv(tmp_path, index=False, encoding=encoding)
            os.replace(tmp_path, path)
        except _WRITE_EXCEPTIONS:
            logger.error("Failed to write data file: %s", path, exc_info=True)
            raise
        finally:
            if tmp_path is not None:
                _remove_quietly(tmp_path)
